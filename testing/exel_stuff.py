# from openpyxl import load_workbook, Workbook
import openpyxl
import random
import ast

max_str_num = 150
#check pats_all exel file if interested

# def rand_pat_from_file():
#     global max_str_num
#     workbook = load_workbook("pats_all.xlsx")
#     sheet = workbook["pats_all"]
#     pat = []
#     while not pat:
#         r = random.randint(2,max_str_num)
#         pat = sheet.cell(row = r, column= 1).value
#     # print(pat)
#     return(pat)

def replace_ints_in_matrix_rev(matrix):
    replacements = {
        70: 10,
        10: 20,
        33: 32,
        30: 33,
        32: 34,
        7040: 41,
        7041: 42,
        1140: 51,
        1141: 52,
        1040: 51,
        1041: 52,
        60: 61,
        63: 62,
        61: 63,
        62: 64,
        7050: 71,
        7052: 72,
        7051: 73,
        7053: 74
    }

    new_matrix = []
    for row in matrix:
        new_row = []
        for element in row:
            if element in replacements:
                new_row.append(replacements[element])
            else:
                new_row.append(element)
        new_matrix.append(new_row)

    return new_matrix

def string_to_list(string_list):
    try:
        # Use ast.literal_eval to safely evaluate the string as a list
        # This is much safer than using 'eval' which could execute arbitrary code
        result_list = ast.literal_eval(string_list)

        # Verify if the result is a list.
        if isinstance(result_list, list):
            return result_list
        else:
            print("Error: The given string does not represent a list.")
            return None

    except:
        pass
# create_file () # if pats all is corrupted, del it and run this def

